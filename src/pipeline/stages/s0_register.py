"""S0 接入与登记:读 manifest + 文件 → 建 doc_versions(REGISTERED / QUARANTINED)。

与生产 §3 一致:manifest 9 必填列(不匹配整批拒收)、SHA-256 精确去重(命中标注关联)、
ULID 双 ID(替代时 logical 继承)、magic number 格式探测(不信扩展名)、隔离路由
(疑似重复 / 密级缺失 / 白名单外)、原件写一次。发文字号/命名仅告警入报告。

s0 是 ingest 入口(非轮询 stage):一次处理整个批次,创建 doc_version 与初始 pipeline_events。
"""

from __future__ import annotations

import hashlib
import io
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from ulid import ULID

from pipeline.index.pg_models import Document, DocVersion, ImportBatch, PipelineEvent
from pipeline.stage_base import StageContext
from pipeline.states import ErrorCode, PipelineState

REQUIRED_COLUMNS = [
    "filename", "title", "doc_number", "issuer", "perm_tag",
    "corpus_type", "biz_domain", "issue_date", "supersedes",
]
WHITELIST_FORMATS = {"docx", "pdf"}


@dataclass
class FileOutcome:
    filename: str
    status: str  # REGISTERED | QUARANTINED | DUPLICATE | MISSING
    doc_version_id: str | None = None
    logical_id: str | None = None
    reason: str = ""
    error_code: str | None = None


@dataclass
class RegisterReport:
    batch_id: str
    accepted: bool  # manifest 9 列校验(整批)
    reject_reason: str = ""
    outcomes: list[FileOutcome] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def counts(self) -> Counter:
        return Counter(o.status for o in self.outcomes)


def detect_format(data: bytes) -> str:
    """magic number 格式探测(不信扩展名):pdf / docx / xlsx / office-other / unknown。"""
    if data[:5] == b"%PDF-":
        return "pdf"
    if data[:4] == b"PK\x03\x04":
        try:
            names = zipfile.ZipFile(io.BytesIO(data)).namelist()
        except zipfile.BadZipFile:
            return "unknown"
        if any(n.startswith("word/") for n in names):
            return "docx"
        if any(n.startswith("xl/") for n in names):
            return "xlsx"
        return "office-other"
    return "unknown"


def _parse_issue_date(value: object) -> date | None:
    """manifest issue_date 归一到 date:openpyxl 日期格给 datetime,文本格给 ISO 字符串。

    空 → None;非空但无法解析 → None(由调用方仅告警入报告,不拒批)。
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):  # datetime 是 date 的子类,须先判
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _read_manifest(path: Path) -> tuple[list, list[dict]]:
    ws = load_workbook(str(path)).active
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None for v in r):
            continue
        rows.append({h: ("" if v is None else v) for h, v in zip(header, r, strict=False)})
    return header, rows


def _find_by_hash(ctx: StageContext, sha: str) -> str | None:
    with ctx.db.session() as s:
        dv = s.scalars(select(DocVersion).where(DocVersion.source_hash == sha)).first()
        return dv.doc_version_id if dv else None


def _suspect_duplicate(ctx: StageContext, title: str, doc_number: str) -> bool:
    if not title or not doc_number:
        return False
    with ctx.db.session() as s:
        q = select(DocVersion).where(
            DocVersion.title == title, DocVersion.doc_number == doc_number
        )
        return s.scalars(q).first() is not None


def _resolve_logical(
    ctx: StageContext, supersedes: str
) -> tuple[str | None, str | None, str | None]:
    """替代声明 → (继承的 logical_id, 被替代 version_id, version_relation);找不到则全 None。"""
    if not supersedes:
        return None, None, None
    with ctx.db.session() as s:
        prior = s.scalars(
            select(DocVersion).where(DocVersion.source_filename == supersedes)
        ).first()
    if prior is None:
        return None, None, None
    return prior.logical_id, prior.doc_version_id, "revise_replace"


def _ensure_batch(
    ctx: StageContext, batch_id: str, batch_dir: Path, manifest_path: Path
) -> None:
    """get-or-create 批次行:同 batch_id 重跑复用既有行(不重插、不动 created_at/report)。

    使 register_batch 对同 batch_id 幂等可重试——中途崩溃后拿同 batch_id 续跑不再撞主键,
    后续每文件 SHA 去重照常命中已登记文件(返回 DUPLICATE,不新建 doc_version → chunk_id 稳定)。
    """
    if ctx.db.get(ImportBatch, batch_id) is not None:
        return
    ctx.db.add(
        ImportBatch(batch_id=batch_id, source_dir=str(batch_dir), manifest_path=str(manifest_path))
    )


def register_batch(
    ctx: StageContext, batch_id: str, batch_dir: Path, manifest_path: Path
) -> RegisterReport:
    header, rows = _read_manifest(Path(manifest_path))
    # SPEC §S0:9 列契约要求列集合精确匹配——缺列/多列均整批拒收(空表头单元格不计为列)。
    header_cols = [c for c in (header or []) if c not in (None, "")]
    missing = [c for c in REQUIRED_COLUMNS if c not in header_cols]
    extra = [c for c in header_cols if c not in REQUIRED_COLUMNS]
    if missing or extra:
        parts = ([f"缺必填列: {missing}"] if missing else []) + (
            [f"多余列: {extra}"] if extra else []
        )
        return RegisterReport(
            batch_id, accepted=False, reject_reason="manifest 列不匹配(" + "; ".join(parts) + ")"
        )

    _ensure_batch(ctx, batch_id, batch_dir, manifest_path)
    report = RegisterReport(batch_id, accepted=True)
    for row in rows:
        if not row.get("filename"):
            continue
        report.outcomes.append(_register_one(ctx, batch_id, Path(batch_dir), row, report))
    return report


def _register_one(
    ctx: StageContext, batch_id: str, batch_dir: Path, row: dict, report: RegisterReport
) -> FileOutcome:
    fn = str(row["filename"])
    path = batch_dir / fn
    if not path.exists():
        return FileOutcome(fn, "MISSING", reason="文件不存在")

    data = path.read_bytes()
    fmt = detect_format(data)
    sha = hashlib.sha256(data).hexdigest()
    corpus = str(row.get("corpus_type") or "")
    perm = str(row.get("perm_tag") or "")
    title = str(row.get("title") or "")
    doc_number = str(row.get("doc_number") or "")

    if not doc_number:
        report.warnings.append(f"{fn}: 发文字号缺失(仅告警)")

    issue_date = _parse_issue_date(row.get("issue_date"))
    if row.get("issue_date") and issue_date is None:
        report.warnings.append(f"{fn}: issue_date 无法解析({row.get('issue_date')!r}),置空(仅告警)")

    dup = _find_by_hash(ctx, sha)
    if dup is not None:  # 精确去重:不重复登记,标注关联
        report.warnings.append(f"{fn}: SHA-256 精确重复,关联 {dup}")
        return FileOutcome(fn, "DUPLICATE", doc_version_id=dup, reason="SHA-256 精确重复")

    # 隔离判定
    reason, ecode = None, None
    if fmt not in WHITELIST_FORMATS:
        reason, ecode = f"格式白名单外({fmt})", ErrorCode.FORMAT_NOT_WHITELISTED.value
    elif not perm:
        reason = "密级缺失"
    elif _suspect_duplicate(ctx, title, doc_number):
        reason = "疑似重复(标题+文号命中,hash 不同)"

    logical_id, supersedes_vid, relation = _resolve_logical(ctx, str(row.get("supersedes") or ""))
    dvid = str(ULID())
    ext = fmt if fmt in WHITELIST_FORMATS else (Path(fn).suffix.lstrip(".") or "bin")
    raw_key = ctx.object_store.put_raw(corpus, batch_id, dvid, ext, data)  # 写一次
    status = PipelineState.QUARANTINED if reason else PipelineState.REGISTERED

    with ctx.db.session() as s:
        if logical_id is None:
            logical_id = str(ULID())
            s.add(Document(logical_id=logical_id, corpus_type=corpus, title=title or None))
            s.flush()  # documents 先落,满足 doc_versions FK
        s.add(
            DocVersion(
                doc_version_id=dvid,
                logical_id=logical_id,
                batch_id=batch_id,
                source_format=fmt,
                source_hash=sha,
                raw_object_key=raw_key,
                source_filename=fn,
                pipeline_status=status.value,
                perm_tag=perm or None,
                biz_domain=str(row.get("biz_domain") or "") or None,
                issuer=str(row.get("issuer") or "") or None,
                doc_number=doc_number or None,
                issue_date=issue_date,
                title=title or None,
                version_relation=relation,
                supersedes_version_id=supersedes_vid,
                last_error_code=ecode,
            )
        )
        s.flush()  # doc_version 先落,满足 pipeline_events FK
        s.add(
            PipelineEvent(
                doc_version_id=dvid,
                from_state=None,
                to_state=status.value,
                error_code=ecode,
                actor=ctx.user,
                detail={"reason": reason} if reason else None,
            )
        )

    return FileOutcome(
        fn, status.value, doc_version_id=dvid, logical_id=logical_id,
        reason=reason or "", error_code=ecode,
    )
