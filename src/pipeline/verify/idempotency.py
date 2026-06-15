"""V5 幂等验证:同批次重复 ingest → chunk_id 集合不变、Milvus num_entities 不变、第二次有事件留痕。

幂等根:s0 SHA-256 精确去重——重复 ingest 同一文件不新建 doc_version(返回 DUPLICATE),既有 doc 写一条
``duplicate_ingest`` pipeline_event(非迁移留痕)。故 chunk_id 集合(确定性 `sha1(dvid|path|seq)`)与
Milvus 实体数都不变。本模块快照「第二次 ingest」前后并比对——**不重嵌入、不需模型**(走 s0 去重路径)。

前置:该批已 ingest 过一次(按 manifest 文件 SHA 找既有 doc_version);未找到则判失败(请先 ingest)。
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from ulid import ULID

from pipeline.index.pg_models import DocVersion, PipelineEvent
from pipeline.stage_base import StageContext
from pipeline.stages.s0_register import register_batch


@dataclass(frozen=True)
class IdempotencyReport:
    passed: bool
    lines: list[str]  # 人读检查项(供 CLI 打印)


def _manifest_filenames(manifest: Path) -> list[str]:
    ws = load_workbook(manifest).active
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    fi = hdr.index("filename")
    return [r[fi] for r in rows[1:] if r[fi]]


def _existing_dvids(ctx: StageContext, batch_dir: Path, manifest: Path) -> list[str]:
    """按 manifest 各文件的 SHA-256 找既有 doc_version(第一次 ingest 登记的)。"""
    out: list[str] = []
    with ctx.db.session() as s:
        for fn in _manifest_filenames(manifest):
            p = batch_dir / fn
            if not p.exists():
                continue
            sha = hashlib.sha256(p.read_bytes()).hexdigest()
            dv = s.scalars(select(DocVersion).where(DocVersion.source_hash == sha)).first()
            if dv is not None:
                out.append(dv.doc_version_id)
    return out


def _chunk_ids(ctx: StageContext, dvid: str) -> frozenset[str]:
    return frozenset(c.chunk_id for c in ctx.db.get_chunks(dvid))


def _dup_event_count(ctx: StageContext, dvids: list[str]) -> int:
    with ctx.db.session() as s:
        evs = s.scalars(
            select(PipelineEvent).where(PipelineEvent.doc_version_id.in_(dvids or [""]))
        )
        return sum(1 for e in evs if (e.detail or {}).get("duplicate_ingest"))


def check_idempotency(ctx: StageContext, batch_dir: Path, manifest: Path) -> IdempotencyReport:
    """快照 → 第二次 register_batch(SHA 去重)→ 再快照 → 比对三项不变量。"""
    dvids = _existing_dvids(ctx, batch_dir, manifest)
    if not dvids:
        return IdempotencyReport(False, ["✗ 未找到已登记文档(请先 demo ingest 该批)"])

    before_chunks = {d: _chunk_ids(ctx, d) for d in dvids}
    before_mcount = {d: ctx.milvus.count(d) for d in dvids}
    before_total = ctx.milvus.count()
    before_dups = _dup_event_count(ctx, dvids)

    register_batch(ctx, str(ULID()), batch_dir, manifest)  # 第二次 ingest → DUPLICATE + 留痕

    after_chunks = {d: _chunk_ids(ctx, d) for d in dvids}
    after_mcount = {d: ctx.milvus.count(d) for d in dvids}
    after_total = ctx.milvus.count()
    after_dups = _dup_event_count(ctx, dvids)

    lines: list[str] = []
    passed = True
    n_chunks = sum(len(v) for v in before_chunks.values())
    if before_chunks == after_chunks:
        lines.append(f"✓ chunk_id 集合不变({n_chunks} 块 / {len(dvids)} 件)")
    else:
        passed = False
        lines.append("✗ chunk_id 集合发生变化(违反确定性/幂等)")
    if before_mcount == after_mcount and before_total == after_total:
        lines.append(f"✓ Milvus num_entities 不变(总 {before_total})")
    else:
        passed = False
        lines.append(f"✗ Milvus 实体数变化(总 {before_total} → {after_total})")
    if after_dups > before_dups:
        lines.append(f"✓ 第二次 ingest 写 duplicate_ingest 留痕(+{after_dups - before_dups})")
    else:
        passed = False
        lines.append("✗ 第二次 ingest 未写 duplicate_ingest 事件")
    return IdempotencyReport(passed, lines)
