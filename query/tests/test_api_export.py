"""T9(SPEC-API §8.1):导出端点。TestClient + fake store;读回 xlsx 验模板 + AI 页脚。

覆盖:xlsx 内容/头、导出权限点 403、消息 404 / 会话不匹配 404、format 非 xlsx→422。
"""

from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from query.api.app import create_app

_PREFIX = "/api/query/v1"
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _svc():
    msg = {
        "conversation_id": "C1", "seq": 2, "content": "已检索到相关制度与案例…",
        "result_json": {
            "route_type": "evidence",
            "citations": [{
                "clause_id": "c1", "doc_title": "《客户适当性管理实施细则》",
                "clause_path": "第三条", "page_start": 7, "status": "effective",
            }],
            "structured": {"cases": {"items": [
                {"title": "某商业银行未评估风险等级案", "regulator": "上海证监局",
                 "penalty_date": "2024-10-17"},
            ]}},
        },
    }
    conv = {"messages": [
        {"role": "user", "seq": 1, "content": "融资融券客户适当性制度依据"},
        {"role": "assistant", "seq": 2, "content": "已检索到相关制度与案例…"},
    ]}
    store = SimpleNamespace(
        get_message=lambda mid: msg if mid == "M2" else None,
        get_conversation=lambda cid: conv if cid == "C1" else None,
    )
    return SimpleNamespace(store=store)


def test_export_xlsx_content_and_headers():
    r = TestClient(create_app(service=_svc())).post(
        f"{_PREFIX}/conversations/C1/messages/M2/export", json={"format": "xlsx"}
    )
    assert r.status_code == 200
    assert r.headers["content-type"] == _XLSX_MEDIA
    assert "attachment" in r.headers["content-disposition"]
    ws = load_workbook(BytesIO(r.content)).active
    vals = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "问题" in vals and "融资融券客户适当性制度依据" in vals   # 配对问句
    assert "路由类型" in vals and "evidence" in vals
    assert "c1" in vals                                          # 依据条款
    assert "某商业银行未评估风险等级案" in vals                    # 相似案例
    assert any("AI 内容标识" in str(v) for v in vals)             # AI 页脚(§9.3)


def test_export_forbidden_403():
    from query.api.auth import Principal, current_principal

    app = create_app(service=_svc())
    app.dependency_overrides[current_principal] = lambda: Principal(can_export=False)
    r = TestClient(app).post(f"{_PREFIX}/conversations/C1/messages/M2/export")
    assert r.status_code == 403 and r.json()["error"]["code"] == "FORBIDDEN"


def test_export_message_not_found_404():
    r = TestClient(create_app(service=_svc())).post(
        f"{_PREFIX}/conversations/C1/messages/NOPE/export"
    )
    assert r.status_code == 404


def test_export_conversation_mismatch_404():
    # M2 属于 C1;在 C9 下请求 → 不匹配 404
    r = TestClient(create_app(service=_svc())).post(
        f"{_PREFIX}/conversations/C9/messages/M2/export"
    )
    assert r.status_code == 404


def test_export_invalid_format_422():
    r = TestClient(create_app(service=_svc())).post(
        f"{_PREFIX}/conversations/C1/messages/M2/export", json={"format": "pdf"}
    )
    assert r.status_code == 422
