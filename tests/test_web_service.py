from pathlib import Path

from openpyxl import load_workbook

from common.pg_models import ReviewQueue
from pipeline.web import app, service


def test_parse_content_type_extracts_boundary():
    main, params = app._parse_content_type("multipart/form-data; boundary=----abc123")
    assert main == "multipart/form-data"
    assert params["boundary"] == "----abc123"


def test_parse_multipart_preserves_binary_and_splits_fields():
    # H2:纯标准库多部件解析(替代已移除的 cgi)。验文本字段 + 文件 + manifest 分流 + **二进制无损**。
    boundary = "----wbk0"
    delim = f"--{boundary}".encode()
    binary = b"%PDF-1.4\r\n\x00\x01\r\nendobj\n"  # 内含 CRLF 且以 \n 结尾(框定剥离不得伤体内字节)
    body = b"\r\n".join([
        delim,
        b'Content-Disposition: form-data; name="corpus_type"',
        b"",
        b"auto",
        delim,
        b'Content-Disposition: form-data; name="files"; filename="a.pdf"',
        b"Content-Type: application/pdf",
        b"",
        binary,
        delim,
        b'Content-Disposition: form-data; name="manifest"; filename="m.xlsx"',
        b"",
        b"XLSXBYTES",
        delim + b"--",
        b"",
    ])
    text, files = app._parse_multipart(body, boundary)
    assert text["corpus_type"] == "auto"
    assert {(fn, fl) for fn, fl, _ in files} == {("files", "a.pdf"), ("manifest", "m.xlsx")}
    pdf = next(data for fn, fl, data in files if fl == "a.pdf")
    assert pdf == binary  # 二进制逐字节无损(含 \r\n、尾 \n)


def test_auto_manifest_uses_contract_columns(tmp_path):
    files = [
        service.UploadedFile("internal_rule.docx", b"x"),
        service.UploadedFile("external_rule.pdf", b"y"),
    ]
    manifest = service._write_auto_manifest(
        tmp_path,
        files,
        corpus_type="auto",
        perm_tag="内部",
        biz_domain="GENERAL",
        issuer="DEMO",
    )

    ws = load_workbook(manifest).active
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == service.REQUIRED_COLUMNS
    assert rows[1][0] == "internal_rule.docx"
    assert rows[1][5] == "P-INT"
    assert rows[2][0] == "external_rule.pdf"
    assert rows[2][5] == "P-EXT"


def test_static_bundle_exists():
    root = Path(service.REPO_ROOT) / "src" / "pipeline" / "web" / "static"
    assert (root / "index.html").exists()
    assert (root / "app.js").exists()
    assert (root / "styles.css").exists()


def test_node_status_treats_open_review_queue_as_waiting():
    queue = ReviewQueue(
        queue_id="01HX0000000000000000000000",
        queue_type="qc_fix",
        doc_version_id="01HY0000000000000000000000",
        status="open",
    )

    assert service._node_status("QC_FAILED", "S2", "S2", [queue]) == "waiting"
    assert service._node_status("QC_FAILED", "S2", "S2", []) == "failed"
